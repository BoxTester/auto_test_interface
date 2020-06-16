# 处理excel中的数据
# 打开excel
# from openpyxl import load_workbook
# wb = load_workbook(r'D:\Python\Project\auto_test_interface\test_data\data.xlsx')
# 定位表单
# sheet = wb['idcard']
# 定位单元格，行列值
# res = sheet.cell(1, 2).value
# print(type(sheet.max_row))  # 最大行
# print(sheet.max_column)  # 最大列
# print(eval(res))  # eval转化为原数据类型

from openpyxl import load_workbook
from project_path import *
from common.tools.do_config import DoConfig


class DoExcel:
    def get_data(self, file_name):
        file_obj = load_workbook(file_name)

        config_data = eval(DoConfig.read_config(conf_path, 'MODE', 'mode')) # mode参数为指定执行用例条数，默认为all（执行所有）
        # 获取数据（以标题为key的字典）
        test_data = []
        for config_item in config_data:
            file_sheet_obj = file_obj[config_item]
            # 获取标题行
            header = []
            for column in range(1, file_sheet_obj.max_column + 1):
                header.append(file_sheet_obj.cell(1, column).value)

            if config_data[config_item] == 'all':
                for row in range(2, file_sheet_obj.max_row + 1):
                    case_data = {}
                    for column in range(1, file_sheet_obj.max_column + 1):
                        case_data[header[column - 1]] = file_sheet_obj.cell(row, column).value
                        case_data['sheet_name'] = config_item
                    test_data.append(case_data)
            else:
                for case_id in config_data[config_item]:
                    case_data = {}
                    for column in range(1, file_sheet_obj.max_column + 1):
                        case_data[header[column - 1]] = file_sheet_obj.cell(case_id + 1, column).value
                        case_data['sheet_name'] = config_item
                    test_data.append(case_data)
        return test_data

    def set_data(self, file_name, sheet_name, row, column, value):
        '''excel写入数据并保存
        file_name：文件名
        sheet_name：表单名
        row、column：行列值（写入的位置）
        value：写入的值'''
        file_obj = load_workbook(file_name)
        file_sheet_obj = file_obj[sheet_name]
        file_sheet_obj.cell(row, column).value = value
        file_obj.save(file_name)


if __name__ == '__main__':
    # config_data = eval(DoConfig.read_config(conf_path,'MODE','mode'))
    # print(config_data)
    # for item in config_data:
    #     print(type(item))
    # sheet_name = 'login'
    # DoExcel(file_name,sheet_name).set_data(1,9,'test_result')
    print(DoExcel().get_data(test_data_path))
